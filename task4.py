import matplotlib.pyplot as plt
from openpyxl import load_workbook
file_path = 'Task4.xlsx'
workbook = load_workbook(file_path)
sheet = workbook.active
values = []
for row in sheet.iter_rows(min_row=2, min_col=3, max_col=3, values_only=True): 
    if row[0] is not None:
        values.append(row[0])
output_txt_path = 'values_task4.txt'
with open(output_txt_path, 'w') as file:
    for value in values:
        file.write(str(value) + '\n')

print(values)
plt.figure(figsize=(10, 6))
plt.bar(range(1, len(values) + 1), values, color='skyblue', edgecolor='black')
plt.title('Values with IDs')
plt.xlabel('ID') 
plt.ylabel('Value')
plt.xticks(range(1, len(values) + 1))
plt.grid(axis='y', linestyle='--', alpha=0.7)
output_image_path = 'histogram_task4.png'
plt.savefig(output_image_path)
plt.close()
print(f"{output_txt_path}")
print(f"{output_image_path}")