from openpyxl import load_workbook
file_path_task1 = "Task1_Task2.xlsx"
output_path_task1 = "results_task1.txt"
workbook_task1 = load_workbook(file_path_task1)
sheet_task1 = workbook_task1.active
values = [cell.value for cell in sheet_task1['B'][1:] if cell.value is not None]
total_sum = sum(values)
average_value = total_sum / len(values)
with open(output_path_task1, 'w', encoding='utf-8') as file:
    file.write(f"Kopējā summa: {total_sum}\n")
    file.write(f"Vidējā vērtība: {average_value:.2f}")
total_sum, average_value


output_path_task2 = "filtered_task2.txt"
filtered_rows = [
    (sheet_task1[f"A{i}"].value, sheet_task1[f"B{i}"].value, sheet_task1[f"C{i}"].value)
    for i in range(2, sheet_task1.max_row + 1)
    if sheet_task1[f"C{i}"].value is not None and sheet_task1[f"C{i}"].value > 50
]
with open(output_path_task2, 'w') as file:
    for row in filtered_rows:
        file.write("\t".join(map(str, row)) + "\n")
len(filtered_rows)