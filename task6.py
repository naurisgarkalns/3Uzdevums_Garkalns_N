from openpyxl import load_workbook
excel_file = 'Task6.xlsx'
txt_file = 'merged_task6.txt'
wb = load_workbook(excel_file)
ws = wb.active
merged_data = []
for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
    name, age = row
    merged_data.append(f"{name}-{age}")
with open(txt_file, 'w') as f:
    f.write("\n".join(merged_data))
print(f"{txt_file}!")