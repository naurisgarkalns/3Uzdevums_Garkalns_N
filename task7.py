from openpyxl import load_workbook
file_path = 'Task7.xlsx'
workbook = load_workbook(file_path)
sheet = workbook.active
old_value = input("Ievadiet vērtību, kuru vēlaties aizvietot no kolonnas E: ")
new_value = input("Ievadiet jauno vērtību, ar kuru aizvietot: ")
updated_data = []
for row in sheet.iter_rows(min_row=2, max_col=5, values_only=False):  
    if row[4].value == old_value:  
        row[4].value = new_value  
    updated_data.append(row[4].value)  
workbook.save(file_path)
output_txt_path = 'updated_task7.txt'
with open(output_txt_path, 'w') as file:
    for item in updated_data:
        file.write(f"{item}\n")
print(f" {output_txt_path}")
